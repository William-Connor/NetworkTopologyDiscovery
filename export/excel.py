"""
LinkExcelExporter — Export inferred topology links and device info
to a formatted Excel (.xlsx) workbook using openpyxl.

Produces two sheets:
  1. 链路清单 — link details with source/target device metadata
  2. 设备清单 — device inventory
"""

from __future__ import annotations

import logging
import os
from datetime import datetime, timezone
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# ── Column definitions for 链路清单 ──────────────────────────────────
LINK_HEADERS: List[str] = [
    "源设备IP",
    "源MAC",
    "源厂商",
    "源设备类型",
    "目标设备IP",
    "目标MAC",
    "目标厂商",
    "目标设备类型",
    "链路类型",
    "跳数",
    "更新时间",
]

LINK_COLUMN_WIDTHS: Dict[int, float] = {
    1: 16,   # 源设备IP
    2: 20,   # 源MAC
    3: 18,   # 源厂商
    4: 12,   # 源设备类型
    5: 16,   # 目标设备IP
    6: 20,   # 目标MAC
    7: 18,   # 目标厂商
    8: 12,   # 目标设备类型
    9: 12,   # 链路类型
    10: 8,   # 跳数
    11: 22,  # 更新时间
}

# ── Column definitions for 设备清单 ──────────────────────────────────
DEVICE_HEADERS: List[str] = [
    "IP",
    "MAC",
    "厂商",
    "设备类型",
    "状态",
    "虚拟",
]

DEVICE_COLUMN_WIDTHS: Dict[int, float] = {
    1: 16,
    2: 20,
    3: 18,
    4: 12,
    5: 10,
    6: 8,
}


class LinkExcelExporter:
    """Export links and devices to a styled Excel workbook."""

    def export(
        self,
        links: List[dict],
        devices: List[dict],
        output_path: str = "output/links.xlsx",
    ) -> str:
        """Write an ``.xlsx`` workbook to *output_path*.

        *links*: list of ``{"source", "target", "type", "hops", "updated"}``
        *devices*: list of device dicts (as produced by ``DeviceClassifier``)

        Returns the absolute path to the saved file.
        """
        # ── Build device lookup ────────────────────────────────────
        dev_by_id: Dict[str, dict] = {}
        for d in devices:
            node_id = self._device_key(d)
            dev_by_id[node_id] = d

        # ── Create workbook ────────────────────────────────────────
        wb = Workbook()

        # ── Sheet 1: 链路清单 ──────────────────────────────────────
        ws1 = wb.active
        ws1.title = "链路清单"
        self._write_link_sheet(ws1, links, dev_by_id)

        # ── Sheet 2: 设备清单 ──────────────────────────────────────
        ws2 = wb.create_sheet("设备清单")
        self._write_device_sheet(ws2, devices)

        # ── Save ───────────────────────────────────────────────────
        out_dir = os.path.dirname(output_path) or "output"
        os.makedirs(out_dir, exist_ok=True)

        abs_path = os.path.abspath(output_path)
        wb.save(abs_path)

        logger.info(
            "LinkExcelExporter: saved → %s (%d links, %d devices)",
            abs_path, len(links), len(devices),
        )
        return abs_path

    # ------------------------------------------------------------------
    # Sheet writers
    # ------------------------------------------------------------------

    def _write_link_sheet(
        self,
        ws,
        links: List[dict],
        dev_by_id: Dict[str, dict],
    ) -> None:
        """Populate the 链路清单 sheet."""
        # Header row
        self._write_header(ws, LINK_HEADERS)

        # Data rows
        for row_idx, link in enumerate(links, start=2):
            src_dev = dev_by_id.get(link.get("source", ""))
            tgt_dev = dev_by_id.get(link.get("target", ""))

            row_data = [
                self._fmt_ip(src_dev),
                self._fmt_mac(src_dev),
                self._fmt_vendor(src_dev),
                self._fmt_type(src_dev),
                self._fmt_ip(tgt_dev),
                self._fmt_mac(tgt_dev),
                self._fmt_vendor(tgt_dev),
                self._fmt_type(tgt_dev),
                link.get("type", "L2_DIRECT"),
                link.get("hops", 1),
                link.get("updated", ""),
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                cell.font = Font(name="Microsoft YaHei", size=10)

            # Alternate-row shading
            if row_idx % 2 == 0:
                fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
                for col_idx in range(1, len(LINK_HEADERS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        # Column widths
        self._set_column_widths(ws, LINK_COLUMN_WIDTHS)

        # Freeze panes & auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(LINK_HEADERS))}{len(links) + 1}"

    def _write_device_sheet(
        self,
        ws,
        devices: List[dict],
    ) -> None:
        """Populate the 设备清单 sheet."""
        # Header row
        self._write_header(ws, DEVICE_HEADERS)

        # Data rows
        for row_idx, dev in enumerate(devices, start=2):
            row_data = [
                dev.get("ip") or "Virtual Switch",
                dev.get("mac") or "-",
                dev.get("vendor", "Unknown"),
                dev.get("type", "endpoint"),
                dev.get("status", "unknown"),
                "是" if dev.get("is_virtual") else "否",
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                cell.font = Font(name="Microsoft YaHei", size=10)

            # Highlight virtual devices
            if dev.get("is_virtual"):
                fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                for col_idx in range(1, len(DEVICE_HEADERS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
            elif row_idx % 2 == 0:
                fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
                for col_idx in range(1, len(DEVICE_HEADERS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        # Column widths
        self._set_column_widths(ws, DEVICE_COLUMN_WIDTHS)

        # Freeze & filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(DEVICE_HEADERS))}{len(devices) + 1}"

    # ------------------------------------------------------------------
    # Formatting helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _fmt_ip(dev: Optional[dict]) -> str:
        if dev is None:
            return "-"
        ip = dev.get("ip")
        if ip is not None:
            return ip
        if dev.get("is_virtual"):
            return "Virtual Switch"
        return "-"

    @staticmethod
    def _fmt_mac(dev: Optional[dict]) -> str:
        if dev is None:
            return "-"
        mac = dev.get("mac")
        if mac is not None:
            return mac
        if dev.get("is_virtual"):
            return "-"
        return "-"

    @staticmethod
    def _fmt_vendor(dev: Optional[dict]) -> str:
        if dev is None:
            return "-"
        vendor = dev.get("vendor", "Unknown")
        if dev.get("is_virtual"):
            return "Virtual"
        return vendor or "Unknown"

    @staticmethod
    def _fmt_type(dev: Optional[dict]) -> str:
        if dev is None:
            return "-"
        return dev.get("type", "unknown")

    @staticmethod
    def _device_key(device: dict) -> str:
        """Return the node identifier used in links (matching GraphBuilder._node_id)."""
        ip = device.get("ip")
        if ip is not None:
            return ip
        if device.get("is_virtual"):
            return "VirtualSwitch"
        mac = device.get("mac")
        if mac:
            return mac
        return "unknown"

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _write_header(ws, headers: List[str]) -> None:
        """Write a styled header row."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    @staticmethod
    def _set_column_widths(ws, width_map: Dict[int, float]) -> None:
        """Apply column widths to the worksheet."""
        for col_idx, width in width_map.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width


# ======================================================================
# Self-test
# ======================================================================
if __name__ == "__main__":
    import argparse

    logging.basicConfig(level=logging.INFO, format="%(message)s")

    parser = argparse.ArgumentParser(description="LinkExcelExporter self-test")
    parser.add_argument(
        "--subnet", default="10.15.117.0/24", help="Subnet to scan"
    )
    parser.add_argument(
        "--gateway", default="10.15.117.1", help="Gateway IP"
    )
    parser.add_argument(
        "--local-ip", default="10.15.117.85", help="Local machine IP"
    )
    parser.add_argument(
        "--interface", "-i", default="eno2", help="ARP interface"
    )
    parser.add_argument(
        "--output", "-o", default="output/links.xlsx", help="Output Excel path"
    )
    args = parser.parse_args()

    from scanner.ping import PingScanner
    from scanner.arp import ARPReader
    from scanner.traceroute import TracerouteScanner
    from topology.classifier import DeviceClassifier
    from topology.linker import LinkInferrer

    print("=" * 60)
    print("  LinkExcelExporter — Self-test")
    print("=" * 60)

    # ── Scan pipeline ─────────────────────────────────────────────
    print("\n[1/4] Ping scan ...")
    ping = PingScanner(timeout=1.0, concurrency=30)
    online = ping.scan_subnet(args.subnet)
    print(f"  Online: {len(online)}")

    print("\n[2/4] ARP cache ...")
    arp_reader = ARPReader()
    arp = arp_reader.read_cache(interface=args.interface)
    print(f"  Entries: {len(arp)}")

    print("\n[3/4] Traceroute + Classify + Link ...")
    tr_targets = ["8.8.8.8", "114.114.114.114", "1.1.1.1"]
    tr = TracerouteScanner(max_hops=15, timeout=3.0, targets=tr_targets)
    tr_results = tr.trace_all()

    classifier = DeviceClassifier()
    devices = classifier.classify(
        online_ips=online, gateway=args.gateway, arp_entries=arp,
        local_ip=args.local_ip, traceroute_results=tr_results,
        traceroute_targets=tr_targets,
    )

    linker = LinkInferrer()
    links = linker.infer(devices, args.subnet, args.gateway, tr_results)
    print(f"  Devices: {len(devices)}, Links: {len(links)}")

    # ── Export Excel ──────────────────────────────────────────────
    print("\n[4/4] Writing Excel ...")
    exporter = LinkExcelExporter()
    result = exporter.export(links, devices, output_path=args.output)

    if result:
        file_size = os.path.getsize(result)
        print(f"  ✅ Excel saved: {result}")
        print(f"  File size: {file_size:,} bytes ({file_size / 1024:.1f} KB)")
        print(f"  Sheets: 链路清单 ({len(links)} rows), 设备清单 ({len(devices)} rows)")
    else:
        print("  ⚠️  No output")
